import io, re
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from .. import models, schemas
from ..database import get_db
from ..auth import get_current_user

router = APIRouter(prefix="/salary", tags=["salary"])


def _parse_note(notes: str | None, key: str) -> float:
    if not notes:
        return 0.0
    m = re.search(rf'{key}:\s*([\d.]+)', notes or '')
    return float(m.group(1)) if m else 0.0


def _build_salary(driver: models.User, reports: list) -> schemas.SalaryOut:
    payments = sum(_parse_note(r.notes, 'Выплата') for r in reports)
    fines    = sum(_parse_note(r.notes, 'Штраф')   for r in reports)
    return schemas.SalaryOut(
        driver_id=driver.id,
        driver_name=driver.full_name,
        total=round(payments - fines, 2),
        base_amount=round(payments, 2),
        bonuses=0.0,
        fines=round(fines, 2),
    )


@router.get("/calculate", response_model=list[schemas.SalaryOut])
def calculate_salary(
    period: str = Query(..., description="YYYY-MM"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    drivers = db.query(models.User).filter(models.User.role == models.UserRole.driver).all()
    result = []
    for driver in drivers:
        reports = db.query(models.Report).filter(
            models.Report.driver_id == driver.id,
            models.Report.shift_date.like(f"{period}%"),
            models.Report.status.in_([
                models.ReportStatus.approved,
                models.ReportStatus.adjusted,
                models.ReportStatus.rejected,
            ]),
        ).all()
        result.append(_build_salary(driver, reports))
    return result


@router.get("/export")
def export_salary(
    period: str = Query(..., description="YYYY-MM"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "openpyxl not installed")

    drivers = db.query(models.User).filter(models.User.role == models.UserRole.driver).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Зарплата {period}"

    header_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["Водитель", "Период", "Начислено", "Штрафы", "Итого"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, driver in enumerate(drivers, 2):
        reports = db.query(models.Report).filter(
            models.Report.driver_id == driver.id,
            models.Report.shift_date.like(f"{period}%"),
            models.Report.status.in_([
                models.ReportStatus.approved,
                models.ReportStatus.adjusted,
                models.ReportStatus.rejected,
            ]),
        ).all()
        s = _build_salary(driver, reports)
        ws.cell(row=row_idx, column=1, value=driver.full_name)
        ws.cell(row=row_idx, column=2, value=period)
        ws.cell(row=row_idx, column=3, value=s.base_amount)
        ws.cell(row=row_idx, column=4, value=s.fines)
        ws.cell(row=row_idx, column=5, value=s.total)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=salary_{period}.xlsx"},
    )
